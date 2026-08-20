from __future__ import annotations

import hashlib
import hmac
import io
import os
from datetime import datetime
from typing import Dict, List

import pandas as pd
import streamlit as st
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from audit_engine import AuditEngine, normalize_transactions
from rules_catalog import AUDIT_GROUPS

APP_NAME = "DeeAU"
APP_TITLE = "DeeAU Transaction Audit Dashboard"

st.set_page_config(
    page_title=APP_TITLE,
    page_icon="assets/deeau_logo.svg",
    layout="wide",
    initial_sidebar_state="expanded",
)


def _get_secret(name: str, default: str) -> str:
    """Read Streamlit secret first, then environment, then local-dev fallback."""
    try:
        value = st.secrets.get(name)
        if value:
            return str(value)
    except Exception:
        pass
    return os.getenv(name, default)


ADMIN_USER = _get_secret("DEE_AU_ADMIN_USER", "DeeLalwani")
ADMIN_PASSWORD = _get_secret("DEE_AU_ADMIN_PASSWORD", "Deepa@2026")


def check_login(username: str, password: str) -> bool:
    return hmac.compare_digest(username, ADMIN_USER) and hmac.compare_digest(password, ADMIN_PASSWORD)


def logo_svg() -> str:
    return """<div class='brand-wrap'>
      <div class='brand-mark'><span>Dee</span><b>AU</b></div>
      <div><div class='brand-name'>DeeAU</div><div class='brand-sub'>Transaction Audit Intelligence</div></div>
    </div>"""


def inject_css() -> None:
    st.markdown(
        """
        <style>
        :root { --deeau-navy:#132238; --deeau-blue:#2563eb; --deeau-cyan:#0891b2; --deeau-bg:#f5f7fb; --deeau-border:#dfe5ef; }
        .stApp { background: linear-gradient(180deg,#f8fafc 0%,#f4f7fb 100%); }
        [data-testid='stSidebar'] { background:#101b2d; }
        [data-testid='stSidebar'] * { color:#f7fafc !important; }
        .brand-wrap { display:flex; align-items:center; gap:14px; margin:4px 0 24px; }
        .brand-mark { width:54px; height:54px; border-radius:16px; display:flex; align-items:center; justify-content:center;
          background:linear-gradient(135deg,#2563eb,#0891b2); color:#fff; font-weight:800; letter-spacing:-1px; box-shadow:0 8px 24px rgba(37,99,235,.25); }
        .brand-mark span { font-size:17px; }
        .brand-mark b { font-size:19px; }
        .brand-name { font-size:30px; line-height:1; font-weight:800; color:var(--deeau-navy); }
        .brand-sub { font-size:12px; color:#64748b; margin-top:5px; }
        .hero { padding:28px 32px; border:1px solid var(--deeau-border); border-radius:22px; background:rgba(255,255,255,.92); box-shadow:0 12px 40px rgba(15,23,42,.06); margin-bottom:20px; }
        .hero h1 { margin:0; color:var(--deeau-navy); font-size:32px; }
        .hero p { color:#64748b; margin:8px 0 0; }
        .metric-card { background:white; border:1px solid var(--deeau-border); border-radius:16px; padding:18px 20px; box-shadow:0 6px 18px rgba(15,23,42,.04); }
        .metric-label { color:#64748b; font-size:12px; font-weight:700; text-transform:uppercase; letter-spacing:.06em; }
        .metric-value { color:var(--deeau-navy); font-size:28px; font-weight:800; margin-top:4px; }
        .login-shell { max-width:520px; margin:7vh auto; }
        .small-note { color:#64748b; font-size:12px; }
        .rule-group { background:white; border:1px solid var(--deeau-border); border-radius:16px; padding:14px 18px; margin-bottom:10px; }
        div[data-testid='stFileUploader'] { border-radius:14px; }
        .stButton>button, .stDownloadButton>button { border-radius:10px; font-weight:700; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def login_page() -> None:
    inject_css()
    st.markdown("<div class='login-shell'>" + logo_svg() + "</div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("Administrator sign-in")
        st.caption("Secure access to the DeeAU transaction audit workspace.")
        with st.form("login_form", clear_on_submit=False):
            username = st.text_input("Login ID", placeholder="Enter your login ID")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            submitted = st.form_submit_button("Sign in", use_container_width=True, type="primary")
        if submitted:
            if check_login(username.strip(), password):
                st.session_state["authenticated"] = True
                st.session_state["username"] = username.strip()
                st.rerun()
            else:
                st.error("Invalid Login ID or Password.")
    st.markdown("<p class='small-note' style='text-align:center;margin-top:16px'>DeeAU • Professional Transaction Audit Workspace</p>", unsafe_allow_html=True)


def metric(label: str, value: str) -> None:
    st.markdown(f"<div class='metric-card'><div class='metric-label'>{label}</div><div class='metric-value'>{value}</div></div>", unsafe_allow_html=True)


def select_rules() -> List[str]:
    st.subheader("2. Select audit procedures")
    st.caption("Select an entire audit group or individual sub-rules. The report will retain the selected rule IDs and remarks per transaction.")
    selected: List[str] = []
    for group in AUDIT_GROUPS:
        group_rules = group["rules"]
        group_key = f"group_{group['id']}"
        group_selected = st.checkbox(
            f"{group['id']} — {group['name']} ({len(group_rules)} rules)",
            key=group_key,
        )
        with st.expander(group["description"], expanded=False):
            for rule in group_rules:
                key = f"rule_{rule['id']}"
                checked = st.checkbox(
                    f"{rule['id']} — {rule['name']}",
                    help=rule["description"],
                    key=key,
                )
                if checked:
                    selected.append(rule["id"])
        if group_selected:
            for rule in group_rules:
                if rule["id"] not in selected:
                    selected.append(rule["id"])
    return selected


def build_excel_report(audit_df: pd.DataFrame, summary_df: pd.DataFrame, rules_df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        audit_df.to_excel(writer, sheet_name="Transaction Audit", index=False)
        summary_df.to_excel(writer, sheet_name="Audit Summary", index=False)
        rules_df.to_excel(writer, sheet_name="Rule Results", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.insert_rows(1, 3)
            ws.freeze_panes = "A5"
            ws["A1"] = "DeeAU | Transaction Audit"
            ws["A1"].font = __import__("openpyxl").styles.Font(bold=True, size=16, color="132238")
            ws["A2"] = "Professional CA Working Paper • Rule-driven exception report"
            ws["A2"].font = __import__("openpyxl").styles.Font(italic=True, size=10, color="64748B")
            logo_path = os.path.join(os.path.dirname(__file__), "assets", "deeau_logo.png")
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 48
                img.height = 48
                ws.add_image(img, "A1")
            ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"
            for column_cells in ws.iter_cols(min_row=4, max_row=ws.max_row):
                max_len = 0
                for cell in column_cells[:250]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)
    return output.getvalue()


def main_app() -> None:
    inject_css()
    st.sidebar.markdown(logo_svg(), unsafe_allow_html=True)
    st.sidebar.success(f"Signed in as {st.session_state.get('username', ADMIN_USER)}")
    if st.sidebar.button("Sign out", use_container_width=True):
        st.session_state.clear()
        st.rerun()
    st.sidebar.markdown("---")
    st.sidebar.caption("DeeAU Transaction Audit")
    st.sidebar.caption("Designed for CA review workflows • Rule-driven • Excel-native")

    st.markdown(
        "<div class='hero'><h1>Transaction Audit Dashboard</h1><p>Upload a transaction workbook, select the required audit procedures, execute the rule engine, and export a CA-ready Excel working paper.</p></div>",
        unsafe_allow_html=True,
    )

    st.subheader("1. Upload transaction workbook")
    uploaded = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"], help="Supported formats: .xlsx and .xls")
    if not uploaded:
        st.info("Upload a transaction workbook to begin. The engine will attempt to identify common transaction columns such as Transaction ID, Date, Amount, Vendor, GSTIN, Tax and Invoice Number.")
        return

    try:
        workbook = pd.ExcelFile(uploaded)
        sheet = st.selectbox("Select transaction sheet", workbook.sheet_names)
        raw_df = workbook.parse(sheet)
    except Exception as exc:
        st.error(f"Unable to read the workbook: {exc}")
        return

    if raw_df.empty:
        st.warning("The selected worksheet is empty.")
        return

    normalized_df, mapping, warnings = normalize_transactions(raw_df)
    if warnings:
        with st.expander("Column mapping notes", expanded=True):
            for warning in warnings:
                st.write(f"• {warning}")

    c1, c2, c3 = st.columns(3)
    with c1: metric("Transactions loaded", f"{len(raw_df):,}")
    with c2: metric("Columns detected", f"{len(mapping):,}")
    with c3: metric("Source sheet", sheet)

    st.markdown("### Detected transaction data")
    st.dataframe(raw_df.head(100), use_container_width=True, height=280)

    selected_rules = select_rules()
    st.caption(f"{len(selected_rules)} audit rule(s) selected.")

    if st.button("Run selected audits", type="primary", use_container_width=True, disabled=not selected_rules):
        with st.spinner("Running DeeAU audit procedures..."):
            engine = AuditEngine(AUDIT_GROUPS)
            result_df, rule_results = engine.run(normalized_df, selected_rules)

            result_df.insert(0, "DeeAU Audit Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            result_df.insert(1, "DeeAU Source Sheet", sheet)

            total_issues = int((result_df["Audit Status"] == "Exception").sum())
            critical = int((result_df["Audit Severity"] == "Critical").sum())
            high = int((result_df["Audit Severity"] == "High").sum())
            medium = int((result_df["Audit Severity"] == "Medium").sum())

            summary_df = pd.DataFrame([
                ["DeeAU", "Transaction Audit", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Source workbook", uploaded.name, ""],
                ["Source sheet", sheet, ""],
                ["Transactions reviewed", len(result_df), ""],
                ["Transactions with exceptions", total_issues, ""],
                ["Critical exceptions", critical, ""],
                ["High exceptions", high, ""],
                ["Medium exceptions", medium, ""],
                ["Rules executed", len(selected_rules), ""],
            ], columns=["Metric", "Value", "Notes"])

            rules_df = pd.DataFrame(rule_results)
            report_bytes = build_excel_report(result_df, summary_df, rules_df)
            st.session_state["audit_result"] = result_df
            st.session_state["audit_summary"] = summary_df
            st.session_state["audit_rules"] = rules_df
            st.session_state["audit_report"] = report_bytes

    if "audit_result" in st.session_state:
        st.markdown("---")
        st.subheader("3. Audit results")
        result_df = st.session_state["audit_result"]
        summary_df = st.session_state["audit_summary"]
        rules_df = st.session_state["audit_rules"]

        e1, e2, e3, e4 = st.columns(4)
        with e1: metric("Exceptions", f"{(result_df['Audit Status'] == 'Exception').sum():,}")
        with e2: metric("Critical", f"{(result_df['Audit Severity'] == 'Critical').sum():,}")
        with e3: metric("High", f"{(result_df['Audit Severity'] == 'High').sum():,}")
        with e4: metric("Rules run", f"{len(rules_df):,}")

        st.dataframe(result_df, use_container_width=True, height=420)
        st.markdown("#### Rule-level summary")
        st.dataframe(rules_df, use_container_width=True, hide_index=True)

        st.download_button(
            "Download transaction_audit.xlsx",
            data=st.session_state["audit_report"],
            file_name="transaction_audit.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        st.caption("Report includes Transaction Audit, Audit Summary and Rule Results sheets. The Transaction Audit sheet retains the original transaction columns plus DeeAU audit status, severity, rule IDs and remarks.")


if __name__ == "__main__":
    if not st.session_state.get("authenticated", False):
        login_page()
    else:
        main_app()
